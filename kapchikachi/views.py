import openpyxl
from django.shortcuts import render, redirect
from django.views import View
import random
from .models import * #모든 모델 상속
from django.views.generic import CreateView #Admin에 띄우면 사용
from openpyxl import Workbook, load_workbook #엑셀DB연결
from .forms import CommentForm #forms의 파일 상속
from django.utils import timezone #시간 설정 패키지
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin #'로그인'에 대한 조건을 담은 클래스
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.http import JsonResponse


#html 연결
def main_html(request): #재윤 기능 test html
    return render(request, 'kapchikachi/index.html')

def thankyou_html(request): #재윤 기능 test html
    return render(request, 'kapchikachi/check.html')

def About_us(request):
    return render(request, 'kapchikachi/about.html')

def notfound_buy(request):
    return render(request, 'kapchikachi/notfound_buy.html')

def Notice_info(request):
    return render(request, 'kapchikachi/Notice_info.html')
def Qna(request):
    return render(request, 'kapchikachi/qna.html') #고동현 추가
#상품목록
def Airpoddog1(request):
    return render(request, 'kapchikachi/products_info1.html')
def Airpoddog2(request):
    return render(request, 'kapchikachi/products_info2.html')
def Airpodcat1(request):
    return render(request, 'kapchikachi/products_info3.html')
def Airpodcat2(request):
    return render(request, 'kapchikachi/products_info4.html')
def KeyringDog(request):
    return render(request, 'kapchikachi/products_info5.html')
def KeyringCat(request):
    return render(request, 'kapchikachi/products_info6.html')
def Products(request):
    return render(request, 'kapchikachi/products.html')

def SelectPayOption(request): #구매 방법 선택(카드결제: 네이버 스토이 이동 / 무통장입금: 우리 구매 페이지 이동)
    return render(request, 'kapchikachi/selectpayoption.html')


###############전역변수 선언###############
login_valid=0 #로그인 값이 유효한지 검사하는 전역변수
count=0 #조회수 체크용

###상품 가격 상품가격 수정(고동현)
price_dog_case1=8900
price_dog_case2=8900
price_dog_keyring=5000
price_cat_case1=8900
price_cat_case2=8900
price_cat_keyring=5000


###상품 원가
cost_dog_case1=6900 #2000원
cost_dag_case2=6900
cost_dog_keyring=4000 #1000원
cost_cat_case1=6900
cost_cat_case2=6900
cost_cat_keyring=4000

#배송비
shipping_fee=2800

#직접수령 장소
pickup_address="한국외국어대학교 글로벌캠퍼스 정문"
direct_pickup="한국외국어대학교 글로벌캠퍼스 정문"

###############기부자들이 로그인 -> 구매 과정###############
@login_required(login_url='common:login')
def CommentCreate(request):

    if request.method == 'POST':
        form = CommentForm(request.POST)

        if form.is_valid():
            comment=form.save(commit=False)
            comment.created=timezone.now()
            comment.author=request.user

            # 필수 정보 체크용(기존 정보 재 이용 시)
            if comment.receiver_name==None or comment.receiver_phone == None or comment.shipping==None or comment.content==None:
                result = {
                    'form': form,
                    'warning': '입력해주세요!'
                }

                return render(request, 'kapchikachi/comment_form.html', context=result)

            #값이 비어있으면 0으로 변경 null 오류 수정
            if comment.dog_case1 == '0':
                comment.dog_case1 =0

            if comment.dog_case2 == '0':
                comment.dog_case2 =0

            if comment.dog_keyring == '0':
                comment.dog_keyring =0

            if comment.cat_case1 == '0':
                comment.cat_case1 =0

            if comment.cat_case2 == '0':
                comment.cat_case2 =0

            if comment.cat_keyring == '0':
                comment.cat_keyring =0
            
            if comment.dog_case1+comment.dog_case2+comment.dog_keyring+comment.cat_case1+comment.cat_case2+comment.cat_keyring == 0:
                result = {
                    'form': form,
                    'warning_product': '상품을 선택해주세요!'
                }

                return render(request, 'kapchikachi/comment_form.html', context=result)

            #구매자 별 총 상품 금액 데이터형 오류 수정(고동현)

            selled_product = int(float(comment.dog_case1))*price_dog_case1 + int(float(comment.dog_case2))*price_dog_case2 + int(float(comment.dog_keyring))*price_dog_keyring + int(float(comment.cat_case1))*price_cat_case1 + int(float(comment.cat_case2))*price_cat_case2 +int(float(comment.cat_keyring))*price_cat_keyring

            #직접수령 시 배송비 x
            #delivery_fee=0

            if comment.shipping=='1': #택배 이용 시 배송비 책정
                # delivery_fee += int(shipping_fee)
                selled_product += 2800

            #배송방법에 따라 상세주소 입력 여부 결정
            if comment.shipping=='1' and comment.detail_address=='' or comment.detail_address==None  :

                result = {
                    'form': form,
                    'warning_address': '배송받을 주소를 입력해주세요!'
                }

                return render(request, 'kapchikachi/comment_form.html', context=result)

            if comment.shipping=='0':
                comment.detail_address=pickup_address

            # if comment.cashreceipts =='1' and (comment.cashreceipts_num =='' or comment.cashreceipts_num ==None):
            #     result = {
            #         'form': form,
            #         'warning_cashreceipts_num': '현금영수증 발급번호를 입력해주세요!'
            #     }
            #
            #     return render(request, 'kapchikachi/comment_form.html', context=result)

            #구매자 별 총 판매 금액
            total_cost=selled_product #+ delivery_fee

            #구매자 총 판매금액
            comment.sell=total_cost

            #'구매 전'임을 DB에 표시
            comment.real_selled=0

            # 상품 준비중 처리
            comment.shipping_state=0


            #구매자 정보 저장
            comment.save()

            return HttpResponseRedirect('/check/')

        else:
            return render(request, 'kapchikachi/comment_form.html', {'form': form})
    else:
        form=CommentForm()
        result = {
            'form': form,
        }
        return render(request, 'kapchikachi/comment_form.html',context=result)


#####기 구매 이력이 있는 회원정보 호출#####
@login_required(login_url='common:login')
def Find_past_info(request):
    this_receiver_name=''
    this_receiver_phone=''

    if request.method != 'POST':
        this_user=request.user.id

        all_data = Comment.objects.all().order_by('-created') #주문 생성 시점 기준 '내림차순'정렬

        user = User.objects.get(id=this_user)
        before_sell=all_data.filter(author_id=user, real_selled=1)
        if int(before_sell.count())<1:
            form = CommentForm()
            result = {
                'form': form,
                'warning_info': '구매이력이 없습니다. 구매 부탁드립니다!',
            }
            return render(request, 'kapchikachi/comment_form.html', context=result)
        else:
            before_info=before_sell.get()

            this_receiver_name=before_info.receiver_name
            this_receiver_phone=before_info.receiver_phone

            form = CommentForm()
            result = {
                'form':form,
                'this_receiver_name':this_receiver_name,
                'this_receiver_phone':this_receiver_phone,
            }
            return render(request, 'kapchikachi/comment_form2.html', context=result)

    ####POST(Submit)받은 경우
    else:
        form = CommentForm(request.POST)

        if form.is_valid():
            comment=form.save(commit=False)
            comment.created=timezone.now()
            comment.author=request.user
            if this_receiver_name != '' and this_receiver_phone != '':
                comment.receiver_name=this_receiver_name
                comment.receiver_phone=this_receiver_phone

            # 필수 정보 체크용(기존 정보 재 이용 시)
            if comment.shipping==None or comment.content==None:
                result = {
                    'form': form,
                    'this_receiver_name':this_receiver_name,
                    'this_receiver_phone':this_receiver_phone,
                    'warnning': '입력해주세요!'
                }

                return render(request, 'kapchikachi/comment_form.html', context=result)

            #값이 비어있으면 0으로 변경
            if comment.dog_case1 == None:
                comment.dog_case1 =0

            if comment.dog_case2 == None:
                comment.dog_case2 =0

            if comment.dog_keyring == None:
                comment.dog_keyring =0

            if comment.cat_case1 == None:
                comment.cat_case1 =0

            if comment.cat_case2 == None:
                comment.cat_case2 =0

            if comment.cat_keyring == None:
                comment.cat_keyring =0

            #구매자 별 총 상품 금액
            selled_product = comment.dog_case1*price_dog_case1 + comment.dog_case2*price_dog_case2 + comment.dog_keyring*price_dog_keyring + comment.cat_case1*price_cat_case1 + comment.cat_case2*price_cat_case2 + comment.cat_keyring*price_cat_keyring

            #직접수령 시 배송비 x
            deliverty_fee=0

            if comment.shipping=='1': #택배 이용 시 배송비 책정
                # deliverty_fee += shipping_fee
                deliverty_fee += 3500

            #배송방법에 따라 상세주소 입력 여부 결정
            if comment.shipping=='1' and comment.detail_address=='' :

                result = {
                    'form': form,
                    'warnning_address': '배송받을 주소를 입력해주세요!'
                }

                return render(request, 'kapchikachi/comment_form2.html', context=result)

            if comment.shipping=='0':
                comment.detail_address=pickup_address

            #구매자 별 총 판매 금액
            total_cost=selled_product + deliverty_fee

            #구매자 총 판매금액
            comment.sell=total_cost+1000

            #'구매 전'임을 DB에 표시
            comment.real_selled=0

            # 상품 준비중 처리
            comment.shipping_state=0

            #구매자 정보 저장
            comment.save()

            return HttpResponseRedirect('/check/')



#####구매전 재확인#####
@login_required(login_url='common:login')
def CheckBeforeSell(request):
    this_user=request.user.id

    all_data = Comment.objects.all().order_by('-created') #주문 생성 시점 기준 '내림차순'정렬

    user = User.objects.get(id=this_user)
    before_sell=all_data.filter(author_id=user, real_selled=0)

    if before_sell.count() ==0: #구매희망 내역이 없는 경우
        return render(request, 'kapchikapchi/not_selected.html') #구매희망 내역이 없다며, 구매창으로 이동할 수 있는 버튼 생성
    ####not_selected.html에 구매창으로 넘길 수 있는 링크 걸기####

    selected_before_sell= before_sell.first()#가장 최신 구매희망 이력만 DB로 불러오기

    #현재 유저가 구매할 정보 확인
    this_purchase_time=selected_before_sell.created #구매 신청 일자



    this_receiver_name =selected_before_sell.receiver_name  # 수취인 성함

    #수취인 번호
    if selected_before_sell.receiver_phone =='':
        this_receiver_phone = '수취인 번호를 다시 입력해주세요.'
    else:
        this_receiver_phone =selected_before_sell.receiver_phone  # 수취인 번호

    #배송방법 선택
    if selected_before_sell.shipping== '0':
        this_shipping = '직접수령'
    if selected_before_sell.shipping== '1':
        this_shipping = '택배배송'
    else:
        this_shipping = '배송방법을 다시 선택하세요.'

    if this_shipping == '1': #배송서비스 이용시 상세주소 입력
        this_receiver_address =selected_before_sell.detail_address  # 수취인 주소(상세주소)
    else:
        this_receiver_address = pickup_address #직접수령 장소

    # 구매상품 확인
    product_type = []  # 구매상품 종류
    if selected_before_sell.dog_case1 > 0:
        product_type.append(f'강아지 에어팟1세대 케이스{selected_before_sell.dog_case1}개')

    if selected_before_sell.dog_case2 > 0:
        product_type.append(f'강아지 에어팟2세대 케이스{selected_before_sell.dog_case2}개')

    if selected_before_sell.dog_keyring > 0:
        product_type.append(f'강아지 키링{selected_before_sell.dog_keyring}개')

    if selected_before_sell.cat_case1 > 0:
        product_type.append(f'고양이 에어팟1세대 케이스{selected_before_sell.cat_case1}개')

    if selected_before_sell.cat_case2 > 0:
        product_type.append(f'고양이 에어팟2세대 케이스{selected_before_sell.cat_case2}개')

    if selected_before_sell.cat_keyring > 0:
        product_type.append(f'고양이 키링{selected_before_sell.cat_keyring}개')

    this_purchase_product='' #상품을 담은 문자열
    for a in range(len(product_type)):
        if a!=0:
            this_purchase_product+=',\n'
        this_purchase_product+=product_type[a]

    #한마디
    this_purchase_comment = selected_before_sell.content

    # 구매금액
    this_purchase_amount = selected_before_sell.sell

    # #현금영수증 발급 여부
    # if selected_before_sell.cashreceipts == '0':
    #     this_cashreceipts = '발급 안함'
    #     this_cashreceipts_num = 'x'
    # else:
    #     this_cashreceipts = '발급'
    #     if selected_before_sell.cashreceipts_num != None:
    #         this_cashreceipts_num = selected_before_sell.cashreceipts_num
    #     else:
    #         this_cashreceipts_num = '현금영수증 발급 번호를 입력해주세요!'


    result={
        'this_purchase_time': this_purchase_time,
        'this_receiver_name':this_receiver_name,
        'this_receiver_phone':this_receiver_phone,
        'this_shipping':this_shipping, #배송방법 -> 동현 프론트추가 요청(재윤)
        'this_receiver_address':this_receiver_address,
        'this_purchase_product':this_purchase_product,
        'this_purchase_comment':this_purchase_comment,
        'this_purchase_amount':this_purchase_amount,
        # 'this_cashreceipts':this_cashreceipts, #현금영수증 -> 동현 프론트추가 요청(재윤)
        # 'this_cashreceipts_num':this_cashreceipts_num #현금영수증번호 -> 동현 프론트추가 요청(재윤)
    }
    return render(request,'kapchikachi/check.html', context=result)

#입금완료 버튼 누를 경우 DB처리 로직
@login_required(login_url='common:login')
def Deposit(request):
    # 사용자 DB설절
    this_user = request.user.id
    user = User.objects.get(id=this_user)

    ##DB추출 및 저장##
    all_data = Comment.objects.all().order_by('-created')  # 주문 생성 시점 기준 '내림차순'정렬

    try: #이미 구매처리가 된 경우엔 F5 버튼으로 새로고침해도 개인 구매내역으로 넘어가게 설정
        personal_data = all_data.filter(author_id=user, real_selled=0) #사용자의 최신 구매요청 정보 추출
        personal_buy=personal_data[0]#.get()
    except:
        return HttpResponseRedirect('/personal/')

    this_pk = personal_buy.pk

    #주문번호를 첫주문을 1000으로 하여 자동연산하도록 지정 --> 서버 실 운영 이후 DB수정 불가능
    # 주문번호 (최초 시작 번호)
    this_order_num = 1000
    this_order_num += int(Order.objects.count())

    this_order_sell = personal_buy.sell

    #구매정보 저장
    this_order_data = Order.objects.create(user_pk = this_pk, user_id = user, order_num = this_order_num, order_sell = this_order_sell, real_selled = 1)
    this_order_data.save()

    #comment DB에서 판매완료 표시
    personal_buy.created = this_order_data.created
    personal_buy.real_selled = 1
    personal_buy.order_num = this_order_num # 주문번호
    personal_buy.deposit_check = 0 # 입금 확인 요청 상태
    personal_buy.save()

    ##사용자가 구매 완료한 시점이므로 구매요청했지만, 실구매가 이루어지지 않은 DB는 모두 삭제
    delete_comment_data = Comment.objects.filter(author_id=user, real_selled=0)
    if delete_comment_data.count() >= 0:
        delete_comment_data.delete()

    # delete_order_data = Order.objects.filter(user_id=user, real_selled=0)
    # if delete_order_data.count() >= 0:
    #     delete_order_data.delete()
    
    # 배송방법 출력
    if personal_buy.shipping == '0':
        this_data_shipping = "직접수령"
    if personal_buy.shipping == '1':
        this_data_shipping = "택배배송"

    # #현금영수증 여부
    # if personal_buy.cashreceipts=='0':
    #     this_data_cashreceipts = "발급안함"
    #     this_data_cashreceipts_num = "x"
    # else:
    #     this_data_cashreceipts = "발급"
    #     if personal_buy.cashreceipts_num != None:
    #         this_data_cashreceipts_num = personal_buy.cashreceipts_num
    #     else:
    #         this_data_cashreceipts_num = "현금영수증 발급을 원하시면 오픈카카오톡으로 문의 부탁드립니다."


    # 관리자가 입금 내역 확인했는 지 여부 출력
    if personal_buy.deposit_check == 0:
        this_deposit_check = "입금확인 전"
    else:
        this_deposit_check = "입금확인 완료"
    
    result={
        'this_created': personal_buy.created, #주문시각
        'this_order_num': this_order_data.order_num, #주문번호
        'this_deposit_name': personal_buy.deposit_name, #예금주 명
        'this_deposit_check' : this_deposit_check, #입금확인 상태
        'this_receiver': personal_buy.receiver_name, #수취인 성함
        'this_receiver_phone': personal_buy.receiver_phone, #수취인 번호
        'this_shipping': this_data_shipping, #배송방법
        'this_address': personal_buy.detail_address, #상세주소
        'this_sell': personal_buy.sell, #판매가격
        'this_comment': personal_buy.content, #구매자의 한마디
        # 'this_data_cashreceipts':this_data_cashreceipts, #현금영수증 발급 여부 -> 동현 프론트 추가요청(재윤)
        # 'this_data_cashreceipts_num':this_data_cashreceipts_num #현금영수증 발급 번호 -> 동현 프론트 추가요청(재윤)
    }

    return render(request, 'kapchikachi/checkbill.html', context=result)

#####개인구매이력 확인창#####
@login_required(login_url='common:login')
def Check_personal_buy(request):
    # 사용자 식별
    this_user = request.user.id
    user = User.objects.get(id=this_user)

    # 사용자의 구매정보 추출(내림차순)
    all_data=Comment.objects.order_by('-created') #내림차순 정렬
    personal_buy= all_data.filter(author_id=user, real_selled=1) #사용자의 구매정보 조회
    personal_buy_num= int(personal_buy.count()) #사용자의 구매횟수 조회

    # 사용자별 주문정보 추출(내림차순)
    all_order=Order.objects.order_by('-created')#내림차순 정렬
    personal_order= all_order.filter(user_id=user, real_selled=1)


    if personal_buy_num <1: #구매이력이 없는 경우
        return HttpResponseRedirect('/notfound_buy/') #urls에서 notfound_buy url과 함수 지정 '구매이력이 없습니다. 캠페인에 동참해주세요.' 문구 출력하는 페이지로 이동
    elif personal_buy_num >3: #구매이력이 3번보다 많은 경우
        personal_buy_num=3
    else: #1~3번 구매한 경우
        pass

    #####구매정보#####(Comment Model DB)
    not_found=[]
    personal_receiver_name=[]
    personal_receiver_phone=[]
    personal_product=[]
    personal_shipping=[]
    personal_address=[] #shipping이 1인 경우 입력/ else인 경우 직접수령
    personal_comment=[]
    personal_shipping_state = []
    personal_order_deposit_name = []
    personal_order_deposit_check = []
    # personal_cashreceipts = []
    # personal_cashreceipts_num = []


    #####주문정보#####(Order Model DB)
    personal_order_created=[]
    personal_order_num=[]
    time_str=[]




    for a in range(0,personal_buy_num): #장고 orm에서 sql의 값을 인덱싱 가능한데, 첫번째 튜플이 index=1일 경우/ 만약 0부터이면 for문 값 바꾸기
        #####구매정보#####
        tmp = personal_buy[a]
        personal_receiver_name.append(f'수령인 명 : {tmp.receiver_name}')
        personal_receiver_phone.append(f'수령인 번호 : {tmp.receiver_phone}')
        if tmp.shipping == 1:
            personal_address.append(f'수령장소 : {tmp.detail_adress}')
            personal_shipping.append('수령방법 : 택배')
        else:
            personal_address.append(f'수령장소 : {pickup_address}')
            personal_shipping.append('수령방법 : 직접수령')
        personal_comment.append(f'응원의 말 : {tmp.content}')
        not_found.append('')

        personal_order_deposit_name.append(f'입금자 명: {tmp.deposit_name}')
        if tmp.deposit_check == 0:
            personal_order_deposit_check.append('주문확인 상태 : 입금확인 전')
        else:
            personal_order_deposit_check.append('주문확인 상태 : 입금확인 완료')

        tmp_product='구매상품 :\n'
        if tmp.dog_case1>=1:
            tmp_product+=f'강아지 에어팟1세대 케이스 {tmp.dog_case1}개\n'
        if tmp.dog_case2>=1:
            tmp_product += f'강아지 에어팟2세대 케이스 {tmp.dog_case2}개\n'
        if tmp.dog_keyring>=1:
            tmp_product += f'강아지 키링 {tmp.dog_keyring}개\n'
        if tmp.cat_case1>=1:
            tmp_product += f'고양이 에어팟1세대 케이스 {tmp.cat_case1}개\n'
        if tmp.cat_case2>=1:
            tmp_product += f'고양이 에어팟2세대 케이스 {tmp.cat_case2}개\n'
        if tmp.cat_keyring>=1:
            tmp_product += f'고양이 키링{tmp.cat_keyring}개\n'
        personal_product.append(tmp_product)

        if tmp.shipping_state =='0':
            personal_shipping_state.append('배송상태 : 준비중')
        elif tmp.shipping_state =='1':
            personal_shipping_state.append('배송상태 : 배송중')
        else:
            personal_shipping_state.append('배송상태 : 배송완료')

        # if tmp.cashreceipts =='0':
        #     personal_cashreceipts.append('현금영수증 : 발급 안함')
        #     personal_cashreceipts_num.append(f'현금영수증 번호 : x')
        # else:
        #     personal_cashreceipts.append('현금영수증 : 발급')
        #     if personal_cashreceipts_num == None:
        #         personal_cashreceipts_num.append(f'현금영수증 번호 : 오픈카카오톡 문의부탁드립니다.')
        #     else:
        #         personal_cashreceipts_num.append(f'현금영수증 번호 : {tmp.cashreceipts_num}')

        #####주문정보#####
        tmp2 = personal_order[a]
        personal_order_created.append(tmp2.created)
        personal_order_num.append(f'주문번호 : {tmp2.order_num}')
        time_str.append('주문시각: ')


    cycle=3-len(personal_receiver_name)
    check_num=len(personal_receiver_name)
    if cycle>0:
        for b in range(1,cycle+1):
            #####구매정보#####
            not_found.append(f'{check_num+b}번째 구매이력이 없습니다.')
            personal_receiver_name.append('')
            personal_receiver_phone.append('')
            personal_product.append('')
            personal_shipping.append('')
            personal_address.append('')  # shipping이 1인 경우 입력/ else인 경우 직접수령
            personal_comment.append('')
            personal_shipping_state.append('')

            personal_order_deposit_name.append('')
            personal_order_deposit_check.append('')
            # personal_cashreceipts.append('')
            # personal_cashreceipts_num.append('')

            #####주문정보#####
            personal_order_created.append('')
            personal_order_num.append('')
            time_str.append('')


    result={
        #####첫번째 주문#####
        'first_not_found' : not_found[0],
        'first_personal_receiver_name' : personal_receiver_name[0],
        'first_personal_receiver_phone' : personal_receiver_phone[0],
        'first_personal_product' : personal_product[0],
        'first_personal_shipping' : personal_shipping[0],
        'first_personal_address' : personal_address[0], # shipping이 1인 경우 입력/ else인 경우 직접수령
        'first_personal_comment' : personal_comment[0],
        'first_personal_order_created' : personal_order_created[0],
        'first_time_str' : time_str[0],
        'first_personal_order_num' : personal_order_num[0],
        'first_personal_shipping_state' : personal_shipping_state[0],
        'first_personal_order_deposit_name':personal_order_deposit_name[0],
        'first_personal_order_deposit_check':personal_order_deposit_check[0],
        # 'first_personal_cashreceipts':personal_cashreceipts[0], #현금영수증 발급 여부 -> 동현 프론트 추가 요청(재윤)
        # 'fisrt_personal_cashreceipts_num':personal_cashreceipts_num[0],#현금영수증 번호 발급 여부 -> 동현 프론트 추가 요청(재윤)

        #####두번째 주문#####
        'second_not_found': not_found[1],
        'second_personal_receiver_name': personal_receiver_name[1],
        'second_personal_receiver_phone': personal_receiver_phone[1],
        'second_personal_product': personal_product[1],
        'second_personal_shipping': personal_shipping[1],
        'second_personal_address': personal_address[1],  # shipping이 1인 경우 입력/ else인 경우 직접수령
        'second_personal_comment': personal_comment[1],
        'second_personal_order_created': personal_order_created[1],
        'second_time_str': time_str[1],
        'second_personal_order_num': personal_order_num[1],
        'second_personal_shipping_state': personal_shipping_state[1],
        'second_personal_order_deposit_name': personal_order_deposit_name[1],
        'second_personal_order_deposit_check': personal_order_deposit_check[1],
        # 'second_personal_cashreceipts':personal_cashreceipts[1],#현금영수증 발급 여부 -> 동현 프론트 추가 요청(재윤)
        # 'second_personal_cashreceipts_num':personal_cashreceipts_num[1],#현금영수증 번호 발급 여부 -> 동현 프론트 추가 요청(재윤)

        #####세번째 주문#####
        'third_not_found': not_found[2],
        'third_personal_receiver_name': personal_receiver_name[2],
        'third_personal_receiver_phone': personal_receiver_phone[2],
        'third_personal_product': personal_product[2],
        'third_personal_shipping': personal_shipping[2],
        'third_personal_address': personal_address[2],  # shipping이 1인 경우 입력/ else인 경우 직접수령
        'third_personal_comment': personal_comment[2],
        'third_personal_order_created': personal_order_created[2],
        'third_time_str': time_str[2],
        'third_personal_order_num': personal_order_num[2],
        'third_personal_shipping_state': personal_shipping_state[2],
        'third_personal_order_deposit_name': personal_order_deposit_name[2],
        'third_personal_order_deposit_check': personal_order_deposit_check[2],
        # 'third_personal_cashreceipts':personal_cashreceipts[2],#현금영수증 발급 여부 -> 동현 프론트 추가 요청(재윤)
        # 'third_personal_cashreceipts_num':personal_cashreceipts_num[2] #현금영수증 번호 발급 여부 -> 동현 프론트 추가 요청(재윤)
    }

    return render(request, 'kapchikachi/personal_buy_list.html', context=result)



###############그래프 표시###############
def GetChart(request):
    #실제 판매된 상품만 객체로 저장
    all_data = Comment.objects.filter(real_selled=1)
    order_data = Order.objects.filter(real_selled=1)

    db_num = int(len(all_data))

    #각 상품 별 판매액
    selled_dog_case1=0
    selled_dog_case2=0
    selled_dog_keyring=0
    selled_cat_case1=0
    selled_cat_case2=0
    selled_cat_keyring=0

    for b in range(db_num):
        selled_dog_case1 += int(all_data[b].dog_case1) * (price_dog_case1 -cost_dog_case1)
        selled_dog_case2 += int(all_data[b].dog_case2) * (price_dog_case2 -cost_dag_case2)
        selled_dog_keyring += int(all_data[b].dog_keyring) * (price_dog_keyring -cost_dog_keyring)
        selled_cat_case1 += int(all_data[b].cat_case1) * (price_cat_case1 -cost_cat_case1)
        selled_cat_case2 += int(all_data[b].cat_case2) * (price_cat_case2 -cost_cat_case2)
        selled_cat_keyring += int(all_data[b].cat_keyring) * (price_cat_keyring -cost_cat_keyring)

        # # tmp= all_data[b].donate_type
        # if tmp =='dog_case1':
        #     selled_dog_case1+=int(all_data[b].sell)
        # elif tmp =='dog_case2':
        #     selled_dog_case2+=int(all_data[b].sell)
        # elif tmp =='dog_keyring':
        #     selled_dog_keyring+=int(all_data[b].sell)
        # elif tmp =='cat_case1':
        #     selled_cat_case1+=int(all_data[b].sell)
        # elif tmp =='cat_case2':
        #     selled_cat_case2+=int(all_data[b].sell)
        # else:
        #     selled_cat_keyring += int(all_data[b].sell)

    # 총 모금액 (판매액 아님)
    total_selled = selled_dog_case1+selled_dog_case2+selled_dog_keyring+selled_cat_case1+selled_cat_case2+selled_cat_keyring

    #유기견 후원액
    money_for_dog=selled_dog_case1+selled_dog_case2+selled_dog_keyring

    #유기묘 후원액
    money_for_cat=selled_cat_case1+selled_cat_case2+selled_cat_keyring

    total_donation=money_for_cat+money_for_dog

    #응원의 글(랜덤 출력) 3가지만 랜덤으로 출력
    ran_num = []
    ran_comment = ['','','']
    ran_comment_time = ['','','']

    if db_num>0:
        ran_num.append(random.randint(1, db_num)-1)
        if db_num>1:
            while True:
                tmp=random.randint(1, db_num)-1
                if (all_data[tmp].content!='' or all_data[tmp].content!=None) and ran_num[0]!= tmp:
                    break
            ran_num.append(tmp)
            if db_num>2:
                while True:
                    tmp = random.randint(1, db_num) - 1
                    if (all_data[tmp].content != '' or all_data[tmp].content != None) and ran_num[1] != tmp and ran_num[0]!= tmp:
                        break
                ran_num.append(tmp)


    if len(ran_num)>0:
        for i in range(len(ran_num)):
            tmp=int(ran_num[i])
            ran_comment_time[i]=order_data[tmp].created
            tmp_name=all_data[tmp].receiver_name

            ran_name= tmp_name[:1] +'**'
            ran_comment[i] = f'{ran_name}: {all_data[tmp].content}'


    result={
        'total_selled':total_selled,
        'dog_case1':selled_dog_case1,
        'dog_case2':selled_dog_case2,
        'dog_keyring':selled_dog_keyring,
        'cat_case1':selled_cat_case1,
        'cat_case2':selled_cat_case2,
        'cat_keyring':selled_cat_keyring,
        'money_for_dog':money_for_dog,
        'money_for_cat':money_for_cat,
        'ran_comment1':ran_comment[0],
        'ran_comment2':ran_comment[1],
        'ran_comment3':ran_comment[2],
        'ran_comment_time1':ran_comment_time[0],
        'ran_comment_time2':ran_comment_time[1],
        'ran_comment_time3':ran_comment_time[2],
        'total_donation':total_donation
    }
    return render(request, 'kapchikachi/index.html',context=result)

#####유저 정보 찾기#####




#################################################################################################

#####결제 시스템#####
@login_required(login_url='common:login')
class OrderCheckoutAjaxView(View):
    def post(self, request, *args, **kwargs):
        if not request.user.is_authenticated():
            return JsonResponse({}, status=401)
        user = request.user

        #사용자 DB설정(내가 추가)
        all_data = Comment.objects.order_by('-created')
        this_user = request.user.id
        user = User.objects.get(id=this_user)

        personal_data = all_data.filter(author_id=user, real_selled=0) #사용자의 최신 구매요청 정보 추출
        personal_buy=personal_data[0]#.get()

        this_pk = personal_buy.pk
        this_id = personal_buy.author_id

        #주문번호를 첫주문을 1000으로 하여 자동연산하도록 지정 --> 서버 실 운영 이후 DB수정 불가능
        this_order_num = 1000
        this_order_num += int(OrderTransaction.objects.count())

        this_order_sell = personal_buy.sell

        # amount = request.POST.get('amount')
        amount = personal_buy.sell
        type = request.POST.get('type')
        try:
            trans = OrderTransaction.objects.create_new(
                user_id=user,
                amount=amount,
                type=type,
                user_pk=this_pk,
                order_num=this_order_num,
                order_sell=this_order_sell,
                real_selled=0,
            )
        except:
            trans = None
        if trans is not None:
            data = {
                "works": True,
                "merchant_id": trans
            }
            return JsonResponse(data)
        else:
            return JsonResponse({}, status=401)

@login_required(login_url='common:login')
class OrderImpAjaxView(View):
    def post(self, request, *args, **kwargs):
        if not request.user.is_authenticated():
            return JsonResponse({}, status=401)
        user = request.user
        merchant_id = request.POST.get('merchant_id')
        imp_id = request.POST.get('imp_id')
        amount = request.POST.get('amount')
        # amount =
        try:
            trans = OrderTransaction.objects.get(
                user_id=user,
                order_id=merchant_id,
                amount=amount
            )
        except:
            trans = None
        if trans is not None:
            trans.transaction_id = imp_id
            trans.success = True
            trans.save()
            data = {
                "works": True
            }
            return JsonResponse(data)
        else:
            return JsonResponse({}, status=401)

def purchase_order(request):
    template = 'kapchikachi/purchase_order.html'
    return render(request, template)




#####결제 중 처리#####
@login_required(login_url='common:login')
def real_sell(request):
    if request.method == 'GET':
        ##### API결제 시스템 #####





        ##DB추출 및 저장##
        all_data = Comment.objects.order_by('-created')

        #사용자 DB설절
        this_user = request.user.id
        user = User.objects.get(id=this_user)

        personal_data = all_data.filter(author_id=user, real_selled=0) #사용자의 최신 구매요청 정보 추출
        personal_buy=personal_data[0]#.get()

        this_pk = personal_buy.pk
        this_id = personal_buy.author_id

        #주문번호를 첫주문을 1000으로 하여 자동연산하도록 지정 --> 서버 실 운영 이후 DB수정 불가능
        this_order_num = 1000
        this_order_num += int(OrderTransaction.objects.count())

        this_order_sell = personal_buy.sell

        #구매정보 저장
        this_bill = OrderTransaction.objects.create(user_pk = this_pk, user_id = user, order_num = this_order_num, order_sell = this_order_sell, real_selled = 0)
        this_bill.save()

        ###### API결제 후 결제 완료 페이지 이동 ######
        return HttpResponseRedirect('/checkbill/')

    else:

        return render(request, 'kapchikachi/payment.html')




############################################
# 만약 결제 완료 창에 render 호출을 할 경우에 새로고침 시 데이터 무한증식 문제 발생시,
# 모든 연산을 바로 위의 함수 '결제 중 처리'에서 API 처리 완료 여부를 if로 확인 후 하기 함수의 연산 적용 예정
#####결제 완료 후 결제완료 창#####
@login_required(login_url='common:login')
def check_bill(request):
    #사용자 식별
    this_user = request.user.id
    user = User.objects.get(id=this_user)

    ##현 함수에서 사용할 DB추출
    unsorted_comment_data = Comment.objects.filter(author_id=user, real_selled=0)
    comment_data = unsorted_comment_data.order_by('-created')
    this_comment_data = comment_data[0]#.get() #현 함수에서 사용할 db

    unsorted_order_data = OrderTransaction.objects.filter(user_id=user, real_selled=0)
    order_data = unsorted_order_data.order_by('-created')
    if order_data.count()<=1:
        this_order_data = order_data
    else:
        this_order_data = order_data[0]#.get() #현 함수에서 사용할 db

    #comment DB에서 판매완료 표시
    this_comment_data.created = this_order_data.created
    this_comment_data.real_selled = 1
    this_comment_data.save()

    #order DB에서 판매완료 표시
    this_order_data.real_selled = 1
    this_order_data.save()

    ##사용자가 구매 완료한 시점이므로 구매요청했지만, 실구매가 이루어지지 않은 DB는 모두 삭제
    delete_comment_data = Comment.objects.filter(author_id=user, real_selled=0)
    if delete_comment_data.count() >= 0:
        delete_comment_data.delete()

    delete_order_data = OrderTransaction.objects.filter(user_id=user, real_selled=0)
    if delete_order_data.count() >= 0:
        delete_order_data.delete()

    result={
        'this_created': this_comment_data.created, #주문시각
        'this_order_num': this_order_data.order_num, #주문번호
        'this_receiver': this_comment_data.receiver_name, #수취인 성함
        'this_receiver_phone': this_comment_data.receiver_phone, #수취인 번호
        'this_shipping': this_comment_data.shipping, #배송방법
        'this_address': this_comment_data.detail_address, #상세주소
        'this_sell': this_comment_data.sell, #판매가격
        'this_comment': this_comment_data.content, #구매자의 한마디
    }

    return render(request, 'kapchikachi/checkbill.html', context=result)



#DB내의 모든 데이터 삭제하기
def DeleteData(request):
    Comment.objects.all().delete() #쿼리의 모든 Comment 데이터 삭제
    # OrderTransaction.objects.all().delete() #쿼리의 모든 Order 데이터 삭제
    return render(request, 'kapchikachi/delete.html')



###############그래프 표시###############
def GetChart2(request):
    #실제 판매된 상품만 객체로 저장
    all_data = Comment.objects.filter(real_selled=1)

    db_num = len(all_data)

    #각 상품 별 판매액
    selled_dog_case1=0
    selled_dog_case2=0
    selled_dog_keyring=0
    selled_cat_case1=0
    selled_cat_case2=0
    selled_cat_keyring=0

    for b in range(db_num):
        selled_dog_case1 += int(all_data[b].dog_case1) * (price_dog_case1 -cost_dog_case1)
        selled_dog_case2 += int(all_data[b].dog_case2) * (price_dog_case2 -cost_dag_case2)
        selled_dog_keyring += int(all_data[b].dog_keyring) * (price_dog_keyring -cost_dog_keyring)
        selled_cat_case1 += int(all_data[b].cat_case1) * (price_cat_case1 -cost_cat_case1)
        selled_cat_case2 += int(all_data[b].cat_case2) * (price_cat_case2 -cost_cat_case2)
        selled_cat_keyring += int(all_data[b].cat_keyring) * (price_cat_keyring -cost_cat_keyring)

        # # tmp= all_data[b].donate_type
        # if tmp =='dog_case1':
        #     selled_dog_case1+=int(all_data[b].sell)
        # elif tmp =='dog_case2':
        #     selled_dog_case2+=int(all_data[b].sell)
        # elif tmp =='dog_keyring':
        #     selled_dog_keyring+=int(all_data[b].sell)
        # elif tmp =='cat_case1':
        #     selled_cat_case1+=int(all_data[b].sell)
        # elif tmp =='cat_case2':
        #     selled_cat_case2+=int(all_data[b].sell)
        # else:
        #     selled_cat_keyring += int(all_data[b].sell)

    # 총 모금액 (판매액 아님)
    total_selled = selled_dog_case1+selled_dog_case2+selled_dog_keyring+selled_cat_case1+selled_cat_case2+selled_cat_keyring

    #유기견 후원액
    money_for_dog=selled_dog_case1+selled_dog_case2+selled_dog_keyring

    #유기묘 후원액
    money_for_cat=selled_cat_case1+selled_cat_case2+selled_cat_keyring

    result={
        'total_selled':total_selled,
        'dog_case1':selled_dog_case1,
        'dog_case2':selled_dog_case2,
        'dog_keyring':selled_dog_keyring,
        'cat_case1':selled_cat_case1,
        'cat_case2':selled_cat_case2,
        'cat_keyring':selled_cat_keyring,
        'money_for_dog':money_for_dog,
        'money_for_cat':money_for_cat,
    }
    return render(request, 'kapchikachi/index.html',context=result)


##########차후 기능 구현#######



#####관리자용 기능 함수#####
#엑셀로 모든 DB반환하기
def GetData(request):
    # 엑셀(DB) 쓰기
    write_wb = openpyxl.load_workbook("C://github/Capstone-Design/donator_list.xlsx")  # 기부자 DB호출

    #sql 내의 모든 데이터 호출
    all_data=Comment.objects.all()
    db_num = len(all_data)

    '''------------------------------------------'''
    # Donator시트 호출
    write_Donator = write_wb['Donator']

    #기부자 명단 작성용 리스트 생성
    all_Donator=[[0,0,0,0,0,0,0,0] for col in range(db_num)] #이상하게 처음 리스트에 sql객체를 담으면 중복으로 담아져서, 첫 for문만 원소를 매번 0으로 새롭게 저장

    for a in range(db_num):
        all_Donator[a][0]=all_data[a].pk #구매자 고유번호
        all_Donator[a][1]=str(all_data[a].author) #구매자 계정명
        all_Donator[a][2]=all_data[a].content #구매자 한줄 응원
        all_Donator[a][3]=all_data[a].receiver_name #수취인 이름
        all_Donator[a][4]=all_data[a].receiver_phone #수취인 번호
        all_Donator[a][5]=int(all_data[a].approximate_address) #수취인 지역(시/도)
        all_Donator[a][6] = all_data[a].detail_address #수취인 상세주소
        all_Donator[a][7]=all_data[a].real_selled #실제 구매 여부
        # all_Donator[a].append(all_data[a].created)

    # Donator시트에 db정보 입력
    for b in range(db_num):
        write_Donator.append(all_Donator[b])

    '''------------------------------------------'''
    #Purchase시트 호출
    write_Purchase = write_wb['Purchase']

    #구매 리스트 작성을 위한 리스트 생성
    all_Purchase=[[] for col in range(db_num)]

    for c in range(db_num):
        all_Purchase[c].append(all_data[c].pk)#사용자 정보

        #사용자 상품별 판매개수
        if all_data[c].dog_case1 >0:
            all_Purchase[c].append(all_data[c].dog_case1)
        else:
            all_Purchase[c].append(0)
        if all_data[c].dog_case2 >0:
            all_Purchase[c].append(all_data[c].dog_case2)
        else:
            all_Purchase[c].append(0)
        if all_data[c].dog_keyring >0:
            all_Purchase[c].append(all_data[c].dog_keyring)
        else:
            all_Purchase[c].append(0)
        if all_data[c].cat_case1 >0:
            all_Purchase[c].append(all_data[c].cat_case1)
        else:
            all_Purchase[c].append(0)
        if all_data[c].cat_case2 >0:
            all_Purchase[c].append(all_data[c].cat_case2)
        else:
            all_Purchase[c].append(0)
        if all_data[c].cat_keyring >0:
            all_Purchase[c].append(all_data[c].cat_keyring)
        else:
            all_Purchase[c].append(0)
        all_Purchase[c].append(int(all_data[c].sell))#사용자당 전체 판매액

    #Purchase시트에 db정보 입력
    for d in range(db_num):
        write_Purchase.append(all_Purchase[d])

    '''------------------------------------------'''

    write_wb.save("C://github/Capstone-Design/donator_list.xlsx")
    return render(request, 'kapchikachi/download.html')


# class CommentCreate1(CreateView):
#     models= Comment
#     fields=[
#         'content', 'created', 'author'
#     ]
#     def get_queryset(self):
#         return Comment.objects.order_by('-created')
#
#     #로그인 조건 추가(sns로그인 시 구매가능하도록 조건 설정)
#     def valid(self, form):
#         current_user = self.request.user
#         if current_user.is_authenticated:
#             form.instance.author = current_user
#             return super(type(self)).valid(form)
#         else:
#             return redirect('/로그인 하라는 안내 페이지/') #동현 페이지 설계 요청 : 로그인이 필요합니다. 페이지
#
#
#     def Write_data(self):
#         #valid가 유효한지 체크한 뒤, author / content를 저장하는 코드
#
#         write_wb= Workbook("C:/github/Capstone-Design/donator_list.xlsx")#기부자 DB호출
#         write_Donator=write_wb['Donator'] #Donator시트 호출
#
#         #Excel 내의 DB설계 시 호출 내용 추가 예정 (미완성)
#         #현재 추가할 사용자를 this_ 수식어로 구분
#
#         this_data=[] #데이터를 넣을 임의의 리스트
#
#         this_pk=Comment.get_donator_pk() #기부자의 pk
#         this_content=Comment.content
#         this_created=Comment.created
#         #sns아이디 축적 시 이용 관련 코드 추가
#
#
#         #기부자(구매자)관련 정보 입력 완료
#         write_Donator.append([this_pk, this_content, this_created])
#         write_Donator.save("C:/github/Capstone-Design/donator_list.xlsx")
#         return

    #사용자 수가 적다면, 전체 DB값을 다 리스트로 불러서, 인덱스 별로 호출하시는 방식으로

    #예시 코드:(전체 내용 불러오기)
    # all_values = [] #모든 데이터 값을 넣을 리스트
    # for row in load_ws.rows:
    #     row_value = []
    #     for cell in row:
    #         row_value.append(cell.value)
    #     all_values.append(row_value)
    #
    # wb = Workbook()


# class ShowChart():

#     엑셀읽어오기
#     load_wb= load_workbook("C:/github/Capstone-Design/donator_list.xlsx")#기부자 DB호출
#     load_ws=load_wb['Donator'] #Donator시트 호출



  # db_num=len(all_data)
    # user_data = [[0, 0, 0, 0, 0] for col in range(db_num)]
    # user_num = []  # 현재 로그인한 유저 식별 번호(전체 데이터 중 현재 유저의 데이터 위치 인덱스 반환)
    # user_purchase_num = 0
    #
    # # 현재 로그인 유저의 데이터 식별기능
    # for a in range(db_num):
    #     if (str(all_data[a].author) == str(request.user.username)) and (int(all_data[a].real_selled)==0):  # 현재 로그인 유저의 데이터 인덱스 번호 조회
    #         user_num.append(a)
    #         user_purchase_num += 1
    #     else:
    #         form = CommentForm()
    #         result = {
    #             'form': form,
    #         }
    #         return render(request, 'kapchikachi/comment_form.html', context=result)

    # 구매내역이 1번이라도 있는 경우
    # if user_purchase_num > 0:
    #     check = 0
    #     user_purchase_time = []
    #     user_product_type = []
    #     user_purchase_amount = []
    #     user_receiver_name = []
    #     user_receiver_phone = []
    #     user_receiver_address = []
    #     # 각각의 구매내역을 result에 추가
    #
    #     for b in range(user_purchase_num):
    #         # 구매정보
    #         # user_name=self.request.user.username #유저이름(식별용)
    #         this_purchase_time = str(all_data[user_num[b]].created)  # 구매일자
    #
    #         this_receiver_name = all_data[user_num[b]].receiver_name  # 수취인 성함
    #         this_receiver_phone = all_data[user_num[b]].receiver_phone  # 수취인 번호
    #         this_receiver_address = all_data[user_num[b]].detail_address  # 수취인 주소(상세주소)
    #
    #         # 구매상품
    #         this_product_type = []  # 구매상품 종류
    #
    #         if all_data[user_num[b]].dog_strap > 0:
    #             this_product_type.append(f'강아지 스트랩 {all_data[user_num[b]].dog_strap}개')
    #         if all_data[user_num[b]].dog_case > 0:
    #             this_product_type.append(f'강아지 케이스 {all_data[user_num[b]].dog_case}개')
    #         if all_data[user_num[b]].cat_strap > 0:
    #             this_product_type.append(f'고양이 스트랩 {all_data[user_num[b]].cat_strap}개')
    #         if all_data[user_num[b]].cat_case > 0:
    #             this_product_type.append(f'고양이 케이스 {all_data[user_num[b]].cat_case}개')
    #
    #         # 구매금액
    #         this_purchase_amount = all_data[user_num[b]].sell
    #
    #         # 반환값
    #         user_purchase_time.append(this_purchase_time)
    #         user_product_type.append(this_product_type)
    #         user_purchase_amount.append(this_purchase_amount)
    #         user_receiver_name.append(this_receiver_name)
    #         user_receiver_phone.append(this_receiver_phone)
    #         user_receiver_address.append(this_receiver_address)
    #
    #         # result={
    #         #     'user_name':user_name,
    #         #     'user_purchase_time':user_purchase_time,
    #         #     'user_product_type':user_product_type,
    #         #     'user_purchase_amount':user_purchase_amount,
    #         #     'user_receiver_name':user_receiver_name,
    #         #     'user_receiver_phone':user_receiver_phone,
    #         #     'user_receiver_address':user_receiver_address,
    #         # }
    #
    #     # total_num=[]
    #     # for i in range(user_purchase_num):
    #     #     total_num.append(str(i))
    #     total_num = int(user_purchase_num)
    #
    #     result = {
    #         'user_purchase_time': user_purchase_time,
    #         'product': user_product_type,
    #         'user_purchase_amount': user_purchase_amount,
    #         'receiver_name': user_receiver_name,
    #         'receiver_phone': user_receiver_phone,
    #         'address': user_receiver_address,
    #         'total_cost': total_num,
    #     }
    #
    #     return render(request, 'check.html', context=result)

